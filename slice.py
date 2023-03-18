from __future__ import annotations

import xml.etree.ElementTree
import pickle
import enum
import json
import os
import io

from fractions import Fraction

from openpyxl_image_loader import SheetImageLoader
from PIL.PngImagePlugin import PngImageFile
from openpyxl import load_workbook

parsed_data = {}

_curses: dict[str, Curse] = {}
_items = {}

def sanitize(x: str) -> str:
    x = x.strip()
    x = x.replace("[purple]", "").replace("[cu]", "")
    return x

class Color(enum.Enum):
    Orange = "topmost"
    Yellow = "top"
    Grey = "middle"
    Red = "bottom"
    Blue = "bottommost"

class Curse:
    def __init__(self, name: str, tier: str, description: str, rarity: str | None, upgrade: str | None, is_gen: str | None):
        self.name = name
        self.tier = int(tier)
        self.description = description
        self.rarity = Fraction(rarity or 1)
        if upgrade is not None:
            upgrade = sanitize(upgrade)
        self._upgrade = upgrade
        self.is_generated = (is_gen == "TRUE")

    @property
    def upgrades_to(self):
        if self._upgrade is not None:
            return _curses[self._upgrade]

    @property
    def display(self):
        return f"{self.name} [{self.tier}] ({self.description})"

class _WithImage:
    def __init__(self, image: PngImageFile):
        self.image = io.BytesIO()
        image.save(self.image, format="png")
        image.close()

class Face(_WithImage):
    def __init__(self, value: str, image: PngImageFile):
        super().__init__(image)
        self.value = value

class Equipment(_WithImage):
    def __init__(self, name: str, value: str, tier: int, image: PngImageFile):
        super().__init__(image)
        self.name = name
        self.value = value
        self.tier = tier

class Hero(_WithImage):
    def __init__(self, name: str, color: Color, tier: int, image: PngImageFile,
                 l: Face, m: Face, t: Face, b: Face, r: Face, rr: Face):
        super().__init__(image)
        self.name = name
        self.color = color
        self.tier = tier
        self.faces = (l, m, t, b, r, rr)

class CurrentRun:
    def __init__(self, data):
        self._data = data["d"]

    @property
    def difficulty(self) -> str:
        return self._data["c"]

    @property
    def curses(self) -> list[str]:
        ret = []
        for curse in self._data["m"]:
            curse = sanitize(curse)
            if curse in _curses:
                ret.append(_curses[curse].display)
            else:
                ret.append(f"<unknown curse {curse!r}>")

        return ret

    @property
    def items(self) -> list[str]:
        ret = []
        for item in self._data["p"]["e"]:
            ret.append(f"{item} [{items[item]['tier']}] ({items[item]['description']})")

        return ret

def get_current_run():
    if "classic" in parsed_data:
        return CurrentRun(parsed_data["classic"])

try:
    from aiohttp.web import Request, Response

    from webpage import router
    from events import add_listener
    from utils import get_req_data
except ModuleNotFoundError: # running as stand-alone module
    pass
else:
    @router.post("/sync/slice")
    async def receive_slice(req: Request):
        data = await get_req_data(req, "data")
        with open(os.path.join("data", "slice-data"), "w") as f:
            f.write(data[0])
        populate(os.path.join("data", "slice-data"))
        run = get_current_run()
        if run is not None:
            return Response(body=pickle.dumps(run.curses))
        return Response()

    @add_listener("setup_init")
    async def _load():
        load()

def populate(path=None) -> bool:
    """Decode the savefile and populate the variables.

    Return True if decoding was successful, False otherwise."""

    if path is None:
        try:
            user = os.environ["USERPROFILE"]
        except KeyError:
            return False

        path = os.path.join(user, ".prefs", "slice-and-dice-2")

    try:
        decoded = xml.etree.ElementTree.parse(path)
    except OSError:
        return False

    root = decoded.getroot()

    for child in root:
        key = child.attrib["key"]
        data = child.text
        if key == "run_history":
            # TODO: non-standard JSON; keys and values aren't quoted
            continue
        parsed_data[key] = json.loads(data)

    return True

def load():
    populate(os.path.join("data", "slice-data"))
    wb = load_workbook("slice_dice_data.xlsx")

    for name, tier, effect, rarity, upgrade, is_gen, *_ in wb["Curses"].iter_rows(3, values_only=True):
        name = sanitize(name)
        _curses[name] = Curse(name, tier, effect, rarity, upgrade, is_gen)

    # TODO
    for file, var in (("items", items),):
        with open(f"{file}.txt") as f:
            cont = False
            for line in f.readlines():
                line = line.replace("\t\t", "\t")[:-1] # remove trailing newline
                if cont:
                    if line.endswith('"'):
                        line = line[:-1]
                        cont = False
                    desc = f"{desc} {line}"
                else:
                    name, tier, desc = line.split("\t")
                    if desc.startswith('"'):
                        desc = desc[1:]
                        cont = True
                if not cont:
                    var[name] = {"tier": tier, "description": desc}
